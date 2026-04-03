"""
PDF and Excel export helpers for all financial reports.
"""
from fpdf import FPDF, XPos, YPos
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


# ── PDF ──────────────────────────────────────────────────────────────────────

HDR_BG   = (30, 42, 68)
HDR_FG   = (255, 255, 255)
ROW_ALT  = (245, 247, 252)
TOTAL_BG = (220, 225, 235)
RED      = (220, 53, 69)
GREEN    = (16, 150, 100)
INDIGO   = (99, 102, 241)


class FinPDF(FPDF):
    def __init__(self, company: str, title: str, subtitle: str = ""):
        super().__init__()
        self._co  = company
        self._ti  = title
        self._sub = subtitle

    def header(self):
        self.set_fill_color(*HDR_BG)
        self.rect(0, 0, 210, 28, "F")
        self.set_text_color(*HDR_FG)
        self.set_font("Helvetica", "B", 13)
        self.set_xy(10, 5)
        self.cell(130, 7, self._co[:50], new_x=XPos.RIGHT, new_y=YPos.TOP)
        self.set_font("Helvetica", "", 9)
        self.set_xy(10, 13)
        self.cell(130, 6, self._ti, new_x=XPos.RIGHT, new_y=YPos.TOP)
        if self._sub:
            self.set_font("Helvetica", "", 7.5)
            self.set_xy(10, 20)
            self.cell(130, 5, self._sub)
        self.set_font("Helvetica", "", 7.5)
        self.set_xy(-75, 14)
        self.cell(65, 5, f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", align="R")
        self.set_text_color(0, 0, 0)
        self.ln(16)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "I", 7)
        self.set_text_color(150, 150, 150)
        self.cell(0, 5, f"FinAI Accounting Platform  |  Page {self.page_no()}", align="C")
        self.set_text_color(0, 0, 0)

    def table_header(self, cols: list, widths: list):
        self.set_fill_color(*HDR_BG)
        self.set_text_color(*HDR_FG)
        self.set_font("Helvetica", "B", 8)
        for label, w in zip(cols, widths):
            self.cell(w, 7, label, border=0, fill=True, align="C",
                      new_x=XPos.RIGHT, new_y=YPos.TOP)
        self.ln(7)
        self.set_text_color(0, 0, 0)

    def table_row(self, values: list, widths: list, fills: list = None,
                  colors: list = None, bold: bool = False, alt: bool = False):
        self.set_font("Helvetica", "B" if bold else "", 8)
        if bold:
            self.set_fill_color(*TOTAL_BG)
        elif alt:
            self.set_fill_color(*ROW_ALT)
        else:
            self.set_fill_color(255, 255, 255)

        for i, (val, w) in enumerate(zip(values, widths)):
            align = fills[i] if fills else "L"
            if colors and colors[i]:
                self.set_text_color(*colors[i])
            self.cell(w, 6.5, str(val), border=0,
                      fill=True, align=align,
                      new_x=XPos.RIGHT, new_y=YPos.TOP)
            if colors and colors[i]:
                self.set_text_color(0, 0, 0)
        self.ln(6.5)

    def section_label(self, label: str):
        self.set_fill_color(*INDIGO)
        self.set_text_color(*HDR_FG)
        self.set_font("Helvetica", "B", 7.5)
        self.cell(0, 6, f"  {label.upper()}", fill=True,
                  new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        self.set_text_color(0, 0, 0)


def _inr(n) -> str:
    n = n or 0
    return f"{n:,.2f}"


# ── Trial Balance PDF ─────────────────────────────────────────────────────────

def trial_balance_pdf(data: dict, company: str) -> bytes:
    pdf = FinPDF(company, "Trial Balance", f"As of: {data.get('as_of','All time')}")
    pdf.add_page()

    groups = {"Assets": [], "Liabilities": [], "Equity": [], "Revenue": [], "Expenses": []}
    for acc in data["accounts"]:
        c = acc["account_code"]
        if c.startswith("1"):   groups["Assets"].append(acc)
        elif c.startswith("2"): groups["Liabilities"].append(acc)
        elif c.startswith("3"): groups["Equity"].append(acc)
        elif c.startswith("4"): groups["Revenue"].append(acc)
        elif c.startswith("5"): groups["Expenses"].append(acc)

    widths = [18, 80, 42, 42]
    pdf.table_header(["Code", "Account Name", "Debit (INR)", "Credit (INR)"], widths)

    alt = False
    for grp, items in groups.items():
        if not items:
            continue
        pdf.section_label(grp)
        for acc in items:
            pdf.table_row(
                [acc["account_code"], acc["account_name"],
                 _inr(acc["debit"]) if acc["debit"] > 0 else "—",
                 _inr(acc["credit"]) if acc["credit"] > 0 else "—"],
                widths,
                fills=["C", "L", "R", "R"],
                colors=[INDIGO, None,
                        RED if acc["debit"] > 0 else None,
                        GREEN if acc["credit"] > 0 else None],
                alt=alt
            )
            alt = not alt

    pdf.ln(2)
    pdf.table_row(["", "GRAND TOTAL", _inr(data["total_debit"]), _inr(data["total_credit"])],
                  widths, fills=["C", "L", "R", "R"],
                  colors=[None, None, RED, GREEN], bold=True)
    return bytes(pdf.output())


# ── P&L PDF ───────────────────────────────────────────────────────────────────

def profit_loss_pdf(data: dict, company: str) -> bytes:
    pdf = FinPDF(company, "Profit & Loss Statement",
                 f"Period: {data.get('from_date','—')} to {data.get('to_date','—')}")
    pdf.add_page()

    widths = [18, 100, 64]

    def section(title, items, total, color):
        pdf.section_label(title)
        alt = False
        for item in items:
            pdf.table_row([item["account_code"], item["account_name"], _inr(item["amount"])],
                          widths, fills=["C", "L", "R"],
                          colors=[INDIGO, None, color], alt=alt)
            alt = not alt
        pdf.table_row(["", f"Total {title}", _inr(total)],
                      widths, fills=["C", "L", "R"],
                      colors=[None, None, color], bold=True)
        pdf.ln(2)

    pdf.table_header(["Code", "Account Name", "Amount (INR)"], widths)
    section("Revenue", data["revenue"], data["total_revenue"], GREEN)
    section("Expenses", data["expenses"], data["total_expenses"], RED)

    net = data["net_profit"]
    pdf.set_fill_color(*(INDIGO if net >= 0 else RED))
    pdf.set_text_color(*HDR_FG)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 10, f"  Net {'Profit' if net >= 0 else 'Loss'}: INR {_inr(abs(net))}",
             fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    return bytes(pdf.output())


# ── Balance Sheet PDF ─────────────────────────────────────────────────────────

def balance_sheet_pdf(data: dict, company: str) -> bytes:
    pdf = FinPDF(company, "Balance Sheet", f"As of: {data.get('as_of','—')}")
    pdf.add_page()

    widths = [18, 100, 64]

    def section(title, items, total, color):
        pdf.section_label(title)
        alt = False
        for item in items:
            pdf.table_row([item["account_code"], item["account_name"], _inr(item["balance"])],
                          widths, fills=["C", "L", "R"],
                          colors=[INDIGO, None, color], alt=alt)
            alt = not alt
        pdf.table_row(["", f"Total {title}", _inr(total)],
                      widths, fills=["C", "L", "R"],
                      colors=[None, None, color], bold=True)
        pdf.ln(2)

    pdf.table_header(["Code", "Account Name", "Balance (INR)"], widths)
    section("Assets", data["assets"], data["total_assets"], INDIGO)
    section("Liabilities", data["liabilities"], data["total_liabilities"], RED)

    eq_items = data["equity"] + [{"account_code": "NET", "account_name": "Net Profit / (Loss)",
                                   "balance": data["net_profit"]}]
    section("Equity", eq_items, data["total_equity"], GREEN)
    return bytes(pdf.output())


# ── GSTR-1 PDF ────────────────────────────────────────────────────────────────

def gstr1_pdf(data: dict, company: str) -> bytes:
    pdf = FinPDF(company, "GSTR-1 — Outward Supplies",
                 f"Period: {data.get('period','—')}")
    pdf.add_page()

    widths = [24, 60, 30, 20, 22, 22, 14]
    pdf.table_header(["Date", "Description", "Taxable Value", "GST%",
                      "CGST", "SGST", "IGST"], widths)

    alt = False
    for txn in data.get("transactions", []):
        pdf.table_row(
            [txn["date"], txn["narration"][:35], _inr(txn["taxable_value"]),
             f"{txn['gst_rate']}%", _inr(txn["cgst"]), _inr(txn["sgst"]),
             _inr(txn["igst"])],
            widths, fills=["C", "L", "R", "C", "R", "R", "R"],
            alt=alt
        )
        alt = not alt

    pdf.ln(3)
    summary = data.get("summary", {})
    pdf.table_row(["", "TOTAL", _inr(summary.get("total_taxable")),
                   "", _inr(summary.get("total_cgst")), _inr(summary.get("total_sgst")),
                   _inr(summary.get("total_igst"))],
                  widths, fills=["C", "L", "R", "C", "R", "R", "R"], bold=True)
    return bytes(pdf.output())


# ── EXCEL HELPERS ─────────────────────────────────────────────────────────────

def _xl_header(ws, row: int, cols: list, col_start: int = 1):
    hdr_fill = PatternFill("solid", fgColor="1E2A44")
    hdr_font = Font(bold=True, color="FFFFFF", size=9)
    for i, val in enumerate(cols):
        c = ws.cell(row=row, column=col_start + i, value=val)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")


def _xl_title(ws, company: str, title: str, subtitle: str, ncols: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=company)
    c.font = Font(bold=True, size=13)
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=title)
    c.font = Font(bold=True, size=11)
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    ws.cell(row=3, column=1, value=subtitle).alignment = Alignment(horizontal="center")


def _xl_section(ws, row: int, label: str, ncols: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=label.upper())
    c.fill = PatternFill("solid", fgColor="6366F1")
    c.font = Font(bold=True, color="FFFFFF", size=8.5)
    return row + 1


def _xl_total(ws, row: int, label: str, *amounts):
    ws.cell(row=row, column=2, value=label).font = Font(bold=True, size=9)
    for i, amt in enumerate(amounts):
        c = ws.cell(row=row, column=3 + i, value=amt)
        c.font = Font(bold=True, size=9)
        c.number_format = "#,##0.00"
        c.fill = PatternFill("solid", fgColor="DCE1EB")


def trial_balance_excel(data: dict, company: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    _xl_title(ws, company, "Trial Balance", f"As of: {data.get('as_of','All time')}", 4)
    _xl_header(ws, 5, ["Account Code", "Account Name", "Debit (INR)", "Credit (INR)"])

    groups = {"Assets": [], "Liabilities": [], "Equity": [], "Revenue": [], "Expenses": []}
    for acc in data["accounts"]:
        c = acc["account_code"]
        if c.startswith("1"):   groups["Assets"].append(acc)
        elif c.startswith("2"): groups["Liabilities"].append(acc)
        elif c.startswith("3"): groups["Equity"].append(acc)
        elif c.startswith("4"): groups["Revenue"].append(acc)
        elif c.startswith("5"): groups["Expenses"].append(acc)

    row = 6
    alt_fill = PatternFill("solid", fgColor="F5F7FC")
    for grp, items in groups.items():
        if not items:
            continue
        row = _xl_section(ws, row, grp, 4)
        for i, acc in enumerate(items):
            ws.cell(row=row, column=1, value=acc["account_code"])
            ws.cell(row=row, column=2, value=acc["account_name"])
            if acc["debit"] > 0:
                c = ws.cell(row=row, column=3, value=acc["debit"])
                c.number_format = "#,##0.00"
                c.font = Font(color="DC3545")
            if acc["credit"] > 0:
                c = ws.cell(row=row, column=4, value=acc["credit"])
                c.number_format = "#,##0.00"
                c.font = Font(color="107050")
            if i % 2:
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = alt_fill
            row += 1

    row += 1
    _xl_total(ws, row, "GRAND TOTAL", data["total_debit"], data["total_credit"])

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def profit_loss_excel(data: dict, company: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L Statement"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 22

    _xl_title(ws, company, "Profit & Loss Statement",
              f"Period: {data.get('from_date')} to {data.get('to_date')}", 3)
    _xl_header(ws, 5, ["Account Code", "Account Name", "Amount (INR)"])

    row = 6
    alt = PatternFill("solid", fgColor="F5F7FC")

    def section(items, label, total, color):
        nonlocal row
        row = _xl_section(ws, row, label, 3)
        for i, item in enumerate(items):
            ws.cell(row=row, column=1, value=item["account_code"])
            ws.cell(row=row, column=2, value=item["account_name"])
            c = ws.cell(row=row, column=3, value=item["amount"])
            c.number_format = "#,##0.00"
            c.font = Font(color=color)
            if i % 2:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = alt
            row += 1
        _xl_total(ws, row, f"Total {label}", total)
        row += 2

    section(data["revenue"], "Revenue", data["total_revenue"], "107050")
    section(data["expenses"], "Expenses", data["total_expenses"], "DC3545")

    net = data["net_profit"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1,
                value=f"Net {'Profit' if net >= 0 else 'Loss'}: INR {net:,.2f}")
    c.font = Font(bold=True, size=11, color="6366F1" if net >= 0 else "DC3545")
    c.fill = PatternFill("solid", fgColor="E8E9FF" if net >= 0 else "FFE4E8")
    c.alignment = Alignment(horizontal="center")

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def balance_sheet_excel(data: dict, company: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 22

    _xl_title(ws, company, "Balance Sheet", f"As of: {data.get('as_of','—')}", 3)
    _xl_header(ws, 5, ["Account Code", "Account Name", "Balance (INR)"])

    row = 6
    alt = PatternFill("solid", fgColor="F5F7FC")

    def section(items, label, total, color):
        nonlocal row
        row = _xl_section(ws, row, label, 3)
        for i, item in enumerate(items):
            ws.cell(row=row, column=1, value=item.get("account_code",""))
            ws.cell(row=row, column=2, value=item["account_name"])
            c = ws.cell(row=row, column=3, value=item["balance"])
            c.number_format = "#,##0.00"
            c.font = Font(color=color)
            if i % 2:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = alt
            row += 1
        _xl_total(ws, row, f"Total {label}", total)
        row += 2

    section(data["assets"], "Assets", data["total_assets"], "6366F1")
    section(data["liabilities"], "Liabilities", data["total_liabilities"], "DC3545")
    eq_items = data["equity"] + [{"account_code": "NET", "account_name": "Net Profit/(Loss)",
                                   "balance": data["net_profit"]}]
    section(eq_items, "Equity", data["total_equity"], "107050")

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def gstr1_excel(data: dict, company: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GSTR-1"
    for col, w in zip("ABCDEFG", [14, 50, 18, 10, 16, 16, 16]):
        ws.column_dimensions[col].width = w

    _xl_title(ws, company, "GSTR-1 — Outward Supplies",
              f"Period: {data.get('period','—')}", 7)
    _xl_header(ws, 5, ["Date", "Description", "Taxable Value", "GST%",
                        "CGST", "SGST", "IGST"])

    alt = PatternFill("solid", fgColor="F5F7FC")
    row = 6
    for i, txn in enumerate(data.get("transactions", [])):
        ws.cell(row=row, column=1, value=txn["date"])
        ws.cell(row=row, column=2, value=txn["narration"])
        for col, val in zip(range(3, 8), [txn["taxable_value"], txn["gst_rate"],
                                           txn["cgst"], txn["sgst"], txn["igst"]]):
            c = ws.cell(row=row, column=col, value=val)
            c.number_format = "#,##0.00" if col != 4 else "0.00"
        if i % 2:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = alt
        row += 1

    s = data.get("summary", {})
    _xl_total(ws, row + 1, "TOTAL",
              s.get("total_taxable", 0), 0,
              s.get("total_cgst", 0), s.get("total_sgst", 0), s.get("total_igst", 0))

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
